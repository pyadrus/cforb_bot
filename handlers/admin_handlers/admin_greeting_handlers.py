import os
import sqlite3

import openpyxl
from aiogram import F
from aiogram import types
from aiogram.filters import Command
from aiogram.filters import StateFilter
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import StatesGroup, State
from aiogram.types import FSInputFile
from loguru import logger

from database.database import reading_from_database
from system.dispatcher import bot, ADMIN_USER_ID
from system.dispatcher import dp
from system.dispatcher import router


@router.message(Command('help'))
async def admin_send_start(message: types.Message, state: FSMContext):
    """–û–±—Ä–∞–±–æ—Ç—á–∏–∫ –∫–æ–º–∞–Ω–¥—ã /start, –æ–Ω –∂–µ –ø–æ—Å—Ç –ø—Ä–∏–≤–µ—Ç—Å—Ç–≤–∏—è üëã"""
    await state.clear()  # –ó–∞–≤–µ—Ä—à–∞–µ–º —Ç–µ–∫—É—â–µ–µ —Å–æ—Å—Ç–æ—è–Ω–∏–µ –º–∞—à–∏–Ω—ã —Å–æ—Å—Ç–æ—è–Ω–∏–π
    """–ê–¥–º–∏–Ω –ø–∞–Ω–µ–ª—å"""
    if message.from_user.id not in ADMIN_USER_ID:
        await message.reply("–£ –≤–∞—Å –Ω–µ—Ç –ø—Ä–∞–≤ –Ω–∞ –≤—ã–ø–æ–ª–Ω–µ–Ω–∏–µ —ç—Ç–æ–π –∫–æ–º–∞–Ω–¥—ã.")
        return
    await message.answer("–ö–æ–º–∞–Ω–¥—ã –∞–¥–º–∏–Ω–∞:\n\n"

                         "<b>–†–µ–¥–∞–∫—Ç–∏—Ä–æ–≤–∞–Ω–∏–µ —Ç–µ–∫—Å—Ç–∞ –∏ –æ—Ç–ø—Ä–∞–≤–∫–∞ —Å–æ–æ–±—â–µ–Ω–∏–π:</b>\n\n"
                         
                         "<b>–†–µ–¥–∞–∫—Ç–∏—Ä–æ–≤–∞–Ω–∏–µ —Ç–µ–∫—Å—Ç–∞:</b>\n"
                         "‚úîÔ∏è /edit_types_of_packaging - –í–∏–¥—ã —É–ø–∞–∫–æ–≤–∫–∏\n"
                         "‚úîÔ∏è /edit_services_and_prices - –£—Å–ª—É–≥–∏ –∏ —Ü–µ–Ω—ã\n"
                         "‚úîÔ∏è /edit_wooden_corners_bag_tape - –î–µ—Ä–µ–≤—è–Ω–Ω—ã–µ —É–≥–æ–ª–∫–∏ + –º–µ—à–æ–∫ + —Å–∫–æ—Ç—á\n"
                         "‚úîÔ∏è /edit_wooden_sheathing_bag_tape - –î–µ—Ä–µ–≤—è–Ω–Ω–∞—è –æ–±—Ä–µ—à–µ—Ç–∫–∞ + –º–µ—à–æ–∫ + —Å–∫–æ—Ç—á\n"
                         "‚úîÔ∏è /edit_order_form - üóí –ë–ª–∞–Ω–∫ –∑–∞–∫–∞–∑–∞\n"
                         "‚úîÔ∏è /edit_bag_tape - –ú–µ—à–æ–∫ + —Å–∫–æ—Ç—á\n"
                         "‚úîÔ∏è /edit_box_bag_tape - –ö–æ—Ä–æ–±–∫–∞ + –º–µ—à–æ–∫ + —Å–∫–æ—Ç—á\n"
                         "‚úîÔ∏è /edit_pallet_crate - –ü–∞–ª–ª–µ—Ç –≤ –æ–±—Ä–µ—à–µ—Ç–∫–µ\n"
                         "‚úîÔ∏è /edit_pallet_with_a_solid_wooden_box - –ü–∞–ª–ª–µ—Ç —Å –≥–ª—É—Ö–∏–º –¥–µ—Ä–µ–≤—è–Ω–Ω—ã–º –∫–æ—Ä–æ–±–æ–º\n"
                         "‚úîÔ∏è /edit_cargo_delivery_prices - –ü—Ä–∞–π—Å—ã –Ω–∞ –¥–æ—Å—Ç–∞–≤–∫—É –ö–∞—Ä–≥–æ\n"
                         "‚úîÔ∏è /edit_goods_redemption_service - –£—Å–ª—É–≥–∞ –í—ã–∫—É–ø–∞ —Ç–æ–≤–∞—Ä–æ–≤\n"
                         "‚úîÔ∏è /edit_product_search_service - –£—Å–ª—É–≥–∞ –ü–æ–∏—Å–∫–∞ —Ç–æ–≤–∞—Ä–æ–≤ (–ø—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª–µ–π –≤ –ö–∏—Ç–∞–µ)\n"
                         "‚úîÔ∏è /edit_supplier_inspection - –ò–Ω—Å–ø–µ–∫—Ü–∏—è –ø–æ—Å—Ç–∞–≤—â–∏–∫–æ–≤ –ø–æ –ø—Ä–æ–≤–∏–Ω—Ü–∏—è–º (–≤—ã–µ–∑–¥ –Ω–∞ –ø—Ä–æ–∏–∑–≤–æ–¥—Å—Ç–≤–æ)\n"
                         "‚úîÔ∏è /edit_wechat_registration_service - –£—Å–ª—É–≥–∞ —Ä–µ–≥–∏—Å—Ç—Ä–∞—Ü–∏–∏ –≤ WeChat\n"
                         "‚úîÔ∏è /edit_purchase_a_supplier_database - –ü—Ä–∏–æ–±—Ä–µ—Å—Ç–∏ –±–∞–∑—É –¥–∞–Ω–Ω—ã—Ö –ø–æ—Å—Ç–∞–≤—â–∏–∫–æ–≤\n"
                         "‚úîÔ∏è /edit_what_payments_await_me - –ö–∞–∫–∏–µ –ø–ª–∞—Ç–µ–∂–∏ –º–µ–Ω—è –æ–∂–∏–¥–∞—é—Ç?\n"
                         "‚úîÔ∏è /edit_how_is_payment_made - –ö–∞–∫ —Å–æ–≤–µ—Ä—à–∞–µ—Ç—Å—è –æ–ø–ª–∞—Ç–∞?\n"
                         "‚úîÔ∏è /edit_self_redemption - üõç –°–∞–º–æ–≤—ã–∫—É–ø\n"
                         "‚úîÔ∏è /edit_reviews - üíå –û—Ç–∑—ã–≤—ã\n"
                         "‚úîÔ∏è /edit_partnership_conditions_for_intermediaries_button - –ü–∞—Ä—Ç–Ω–µ—Ä—Å–∫–∏–µ —É—Å–ª–æ–≤–∏—è –¥–ª—è –ø–æ—Å—Ä–µ–¥–Ω–∏–∫–æ–≤"
                         "‚úîÔ∏è /edit_main_menu - —Ç–µ–∫—Å—Ç –º–µ–Ω—é –±–æ—Ç–∞\n"
                         "‚úîÔ∏è /edit_white_cargo_gte - –ë–µ–ª–∞—è –¥–æ—Å—Ç–∞–≤–∫–∞ –≥—Ä—É–∑–æ–≤ —Å –ì–¢–î\n\n"

                         "<b>–ü–æ–ª—É—á–µ–Ω–∏–µ –¥–∞–Ω–Ω—ã—Ö:</b>\n"
                         "‚úîÔ∏è /get_a_list_of_users_registered_in_the_bot - –ü–æ–ª—É—á–µ–Ω–∏–µ —Å–ø–∏—Å–∫–∞ –∑–∞—Ä–µ–≥–∏—Å—Ç—Ä–∏—Ä–æ–≤–∞–Ω–Ω—ã—Ö "
                         "–ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–∏ÃÜ\n"
                         "‚úîÔ∏è /get_users_who_launched_the_bot - –ü–æ–ª—É—á–µ–Ω–∏–µ –¥–∞–Ω–Ω—ã—Ö –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π, –∑–∞–ø—É—Å–∫–∞—é—â–∏—Ö –±–æ—Ç–∞\n\n"

                         "<b>–û—Ç–ø—Ä–∞–≤–∫–∞ —Å–æ–æ–±—â–µ–Ω–∏–π:</b>\n"
                         "‚úîÔ∏è /send_an_image_to_bot_users - –û—Ç–ø—Ä–∞–≤–∫–∞ –∏–∑–æ–±—Ä–∞–∂–µ–Ω–∏—è —á–µ—Ä–µ–∑ –±–æ—Ç–∞ + —Ç–µ–∫—Å—Ç\n"
                         "‚úîÔ∏è /send_a_message_to_bot_users - –û—Ç–ø—Ä–∞–≤–∫–∞ —Ç–µ–∫—Å—Ç–∞ —á–µ—Ä–µ–∑ –±–æ—Ç–∞\n\n"

                         "<b>–ó–∞–º–µ–Ω–∞ –∏–∑–æ–±—Ä–∞–∂–µ–Ω–∏—è –ø–æ—Å—Ç–æ–≤ –∏ —Ñ–∞–π–ª–æ–≤:</b>\n"
                         "‚úîÔ∏è /greeting_photo - –ü–æ—Å—Ç –ø—Ä–∏–≤–µ—Ç—Å—Ç–≤–∏–µ\n"
                         "‚úîÔ∏è /get_price_lists_file - –ó–∞–º–µ–Ω–∞ –ü—Ä–µ–π—Å–∫—É—Ä–∞–Ω—Ç–∞ CFORB.xlsx"
                         "‚úîÔ∏è /services_and_prices_photo - –£—Å–ª—É–≥–∏ –∏ —Ü–µ–Ω—ã\n"
                         "‚úîÔ∏è /white_cargo_gte_photo - –ë–µ–ª–∞—è –¥–æ—Å—Ç–∞–≤–∫–∞ –≥—Ä—É–∑–æ–≤ —Å –ì–¢–î "
                         "‚úîÔ∏è /types_of_packaging_photo - –í–∏–¥—ã —É–ø–∞–∫–æ–≤–∫–∏\n\n"

                         "/start - –Ω–∞—á–∞–ª—å–Ω–æ–µ –º–µ–Ω—é\n", parse_mode="HTML")


# –§—É–Ω–∫—Ü–∏—è –¥–ª—è —Å–æ–∑–¥–∞–Ω–∏—è —Ñ–∞–π–ª–∞ Excel —Å –¥–∞–Ω–Ω—ã–º–∏ –∑–∞–∫–∞–∑–æ–≤
def create_excel_file(orders):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    # –ó–∞–≥–æ–ª–æ–≤–∫–∏ —Å—Ç–æ–ª–±—Ü–æ–≤
    sheet['A1'] = 'ID –∞–∫–∫–∞—É–Ω—Ç–∞ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è'
    sheet['B1'] = '–ò–º—è'
    sheet['C1'] = '–§–∞–º–∏–ª–∏—è'
    sheet['D1'] = '–ì–æ—Ä–æ–¥'
    sheet['E1'] = '–ù–æ–º–µ—Ä —Ç–µ–ª–µ—Ñ–æ–Ω–∞'
    sheet['F1'] = '–î–∞—Ç–∞ —Ä–µ–≥–∏—Å—Ç—Ä–∞—Ü–∏–∏'
    # –ó–∞–ø–æ–ª–Ω–µ–Ω–∏–µ –¥–∞–Ω–Ω—ã–º–∏ –∑–∞–∫–∞–∑–æ–≤
    for index, order in enumerate(orders, start=2):
        sheet.cell(row=index, column=1).value = order[0]  # ID –∞–∫–∫–∞—É–Ω—Ç–∞ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è
        sheet.cell(row=index, column=2).value = order[1]  # –ò–º—è
        sheet.cell(row=index, column=3).value = order[2]  # –§–∞–º–∏–ª–∏—è
        sheet.cell(row=index, column=4).value = order[3]  # –ì–æ—Ä–æ–¥
        sheet.cell(row=index, column=5).value = order[4]  # –ù–æ–º–µ—Ä —Ç–µ–ª–µ—Ñ–æ–Ω–∞
        sheet.cell(row=index, column=6).value = order[5]  # –î–∞—Ç–∞ —Ä–µ–≥–∏—Å—Ç—Ä–∞—Ü–∏–∏

    return workbook


@router.message(Command('get_a_list_of_users_registered_in_the_bot'))
async def export_data(message: types.Message, state: FSMContext):
    """–ü–æ–ª—É—á–µ–Ω–∏–µ —Å–ø–∏—Å–∫–∞ –∑–∞—Ä–µ–≥–∏—Å—Ç—Ä–∏—Ä–æ–≤–∞–Ω–Ω—ã—Ö –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–∏ÃÜ"""
    await state.clear()  # –ó–∞–≤–µ—Ä—à–∞–µ–º —Ç–µ–∫—É—â–µ–µ —Å–æ—Å—Ç–æ—è–Ω–∏–µ –º–∞—à–∏–Ω—ã —Å–æ—Å—Ç–æ—è–Ω–∏–π
    try:
        if message.from_user.id not in [535185511, 301634256]:
            await message.reply('–£ –≤–∞—Å –Ω–µ—Ç –¥–æ—Å—Ç—É–ø–∞ –∫ —ç—Ç–æ–π –∫–æ–º–∞–Ω–¥–µ.')
            return
        orders = reading_from_database()
        # –°–æ–∑–¥–∞–Ω–∏–µ —Ñ–∞–π–ª–∞ Excel
        workbook = create_excel_file(orders)
        filename = '–ó–∞—Ä–µ–≥–∏—Å—Ç—Ä–∏—Ä–æ–≤–∞–Ω–Ω—ã–µ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–∏ –≤ –±–æ—Ç–µ.xlsx'
        workbook.save(filename)  # –°–æ—Ö—Ä–∞–Ω–µ–Ω–∏–µ —Ñ–∞–π–ª–∞
        text = ("–î–∞–Ω–Ω—ã–µ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π –∑–∞—Ä–µ–≥–∏—Å—Ç—Ä–∏—Ä–æ–≤–∞–Ω–Ω—ã—Ö –≤ –±–æ—Ç–µ\n\n"
                "–î–ª—è –≤–æ–∑–≤—Ä–∞—Ç–∞ –≤ –Ω–∞—á–∞–ª—å–Ω–æ–µ –º–µ–Ω—é –Ω–∞–∂–º–∏ –Ω–∞ /start –∏–ª–∏ /help")
        file = FSInputFile(filename)
        await bot.send_document(message.from_user.id, document=file, caption=text)  # –û—Ç–ø—Ä–∞–≤–∫–∞ —Ñ–∞–π–ª–∞ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—é
        os.remove(filename)  # –£–¥–∞–ª–µ–Ω–∏–µ —Ñ–∞–π–ª–∞
    except Exception as e:
        logger.error(e)


def create_excel_file_start(orders):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    # –ó–∞–≥–æ–ª–æ–≤–∫–∏ —Å—Ç–æ–ª–±—Ü–æ–≤
    sheet['A1'] = 'ID –∞–∫–∫–∞—É–Ω—Ç–∞ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è'
    sheet['B1'] = 'username'
    sheet['C1'] = '–ò–º—è'
    sheet['D1'] = '–§–∞–º–∏–ª–∏—è'
    sheet['E1'] = '–î–∞—Ç–∞ –∑–∞–ø—É—Å–∫–∞ –±–æ—Ç–∞'
    # –ó–∞–ø–æ–ª–Ω–µ–Ω–∏–µ –¥–∞–Ω–Ω—ã–º–∏ –∑–∞–∫–∞–∑–æ–≤
    for index, order in enumerate(orders, start=2):
        sheet.cell(row=index, column=1).value = order[0]  # ID –∞–∫–∫–∞—É–Ω—Ç–∞ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è
        sheet.cell(row=index, column=2).value = order[1]  # username
        sheet.cell(row=index, column=3).value = order[2]  # –ò–º—è
        sheet.cell(row=index, column=4).value = order[3]  # –§–∞–º–∏–ª–∏—è
        sheet.cell(row=index, column=5).value = order[4]  # –î–∞—Ç–∞ –∑–∞–ø—É—Å–∫–∞ –±–æ—Ç–∞

    return workbook


@router.message(Command("get_users_who_launched_the_bot"))
async def get_users_who_launched_the_bot(message: types.Message, state: FSMContext):
    """–ü–æ–ª—É—á–µ–Ω–∏–µ –¥–∞–Ω–Ω—ã—Ö –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π, –∑–∞–ø—É—Å–∫–∞—é—â–∏—Ö –±–æ—Ç–∞"""
    await state.clear()  # –ó–∞–≤–µ—Ä—à–∞–µ–º —Ç–µ–∫—É—â–µ–µ —Å–æ—Å—Ç–æ—è–Ω–∏–µ –º–∞—à–∏–Ω—ã —Å–æ—Å—Ç–æ—è–Ω–∏–π
    try:
        if message.from_user.id not in [535185511, 301634256]:
            await message.reply('–£ –≤–∞—Å –Ω–µ—Ç –¥–æ—Å—Ç—É–ø–∞ –∫ —ç—Ç–æ–π –∫–æ–º–∞–Ω–¥–µ.')
            return
        orders = reading_from_database()
        workbook = create_excel_file_start(orders)  # –°–æ–∑–¥–∞–Ω–∏–µ —Ñ–∞–π–ª–∞ Excel
        filename = '–î–∞–Ω–Ω—ã–µ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π –∑–∞–ø—É—Å—Ç–∏–≤—à–∏—Ö –±–æ—Ç–∞.xlsx'
        workbook.save(filename)  # –°–æ—Ö—Ä–∞–Ω–µ–Ω–∏–µ —Ñ–∞–π–ª–∞
        file = FSInputFile(filename)
        text = ("–î–∞–Ω–Ω—ã–µ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π –∑–∞—Ä–µ–≥–∏—Å—Ç—Ä–∏—Ä–æ–≤–∞–Ω–Ω—ã—Ö –≤ –±–æ—Ç–µ\n\n"
                "–î–ª—è –≤–æ–∑–≤—Ä–∞—Ç–∞ –≤ –Ω–∞—á–∞–ª—å–Ω–æ–µ –º–µ–Ω—é –Ω–∞–∂–º–∏ –Ω–∞ /start –∏–ª–∏ /help")
        await bot.send_document(message.from_user.id, document=file, caption=text)  # –û—Ç–ø—Ä–∞–≤–∫–∞ —Ñ–∞–π–ª–∞ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—é
        os.remove(filename)  # –£–¥–∞–ª–µ–Ω–∏–µ —Ñ–∞–π–ª–∞
    except Exception as e:
        logger.error(e)


class MyStates(StatesGroup):
    waiting_for_message = State()
    waiting_for_image = State()
    waiting_for_caption = State()


@router.callback_query(F.data == 'send_an_image_to_bot_users')
async def send_an_image_to_bot_users(message: types.Message, state: FSMContext):
    """–ó–∞–ø—Ä–∞—à–∏–≤–∞–µ—Ç –∏–∑–æ–±—Ä–∞–∂–µ–Ω–∏–µ —É –∞–¥–º–∏–Ω–∏—Å—Ç—Ä–∞—Ç–æ—Ä–∞"""
    await state.clear()  # –ó–∞–≤–µ—Ä—à–∞–µ–º —Ç–µ–∫—É—â–µ–µ —Å–æ—Å—Ç–æ—è–Ω–∏–µ –º–∞—à–∏–Ω—ã —Å–æ—Å—Ç–æ—è–Ω–∏–π
    try:
        if message.from_user.id not in [535185511, 301634256]:
            await message.reply('–£ –≤–∞—Å –Ω–µ—Ç –¥–æ—Å—Ç—É–ø–∞ –∫ —ç—Ç–æ–π –∫–æ–º–∞–Ω–¥–µ.')
            return
        await bot.send_message(message.from_user.id, text="–ó–∞–≥—Ä—É–∑–∏—Ç–µ –∏–∑–æ–±—Ä–∞–∂–µ–Ω–∏–µ –¥–ª—è —Ä–∞—Å—Å—ã–ª–∫–∏:")
        await state.set_state(MyStates.waiting_for_image)  # –£—Å—Ç–∞–Ω–∞–≤–ª–∏–≤–∞–µ–º —Å–æ—Å—Ç–æ—è–Ω–∏–µ "–æ–∂–∏–¥–∞–Ω–∏–µ –∏–∑–æ–±—Ä–∞–∂–µ–Ω–∏—è"
    except Exception as e:
        logger.error(e)


@router.message(StateFilter(MyStates.waiting_for_image))
async def process_send_image(message: types.Message, state: FSMContext):
    """
    –≠—Ç–æ—Ç —Ö–µ–Ω–¥–ª–µ—Ä –±—É–¥–µ—Ç –∂–¥–∞—Ç—å –∑–∞–≥—Ä—É–∂–µ–Ω–Ω–æ–≥–æ –∏–∑–æ–±—Ä–∞–∂–µ–Ω–∏—è –∏ –ø–µ—Ä–µ—Ö–æ–¥–∏—Ç—å –≤ —Å–æ—Å—Ç–æ—è–Ω–∏–µ "–æ–∂–∏–¥–∞–Ω–∏–µ –ø–æ–¥–ø–∏—Å–∏"
    """

    await state.update_data(photo=message.photo[-1].file_id)
    await bot.send_message(message.from_user.id, text="–í–≤–µ–¥–∏—Ç–µ –ø–æ–¥–ø–∏—Å—å –∫ –∏–∑–æ–±—Ä–∞–∂–µ–Ω–∏—é:")
    await state.set_state(MyStates.waiting_for_caption)


@router.message(StateFilter(MyStates.waiting_for_caption))
async def process_send_image_with_caption(message: types.Message, state: FSMContext):
    """
    –≠—Ç–æ—Ç —Ö–µ–Ω–¥–ª–µ—Ä –±—É–¥–µ—Ç –∂–¥–∞—Ç—å –≤–≤–µ–¥–µ–Ω–Ω–æ–π –ø–æ–¥–ø–∏—Å–∏ –∏ –≤—ã–ø–æ–ª–Ω—è—Ç—å —Ä–∞—Å—Å—ã–ª–∫—É
    """
    state_data = await state.get_data()  # –ü–æ–ª—É—á–∏—Ç—å –¥–∞–Ω–Ω—ã–µ –æ —Å–æ—Å—Ç–æ—è–Ω–∏–∏
    state_data['caption'] = message.text  # –°–æ—Ö—Ä–∞–Ω–∏—Ç–µ –∑–∞–≥–æ–ª–æ–≤–æ–∫ –≤ –¥–∞–Ω–Ω—ã—Ö —Å–æ—Å—Ç–æ—è–Ω–∏—è
    photo = state_data.get('photo')  # –ü–æ–ª—É—á–∏—Ç–µ —Ñ–æ—Ç–æ–≥—Ä–∞—Ñ–∏—é –∏ –ø–æ–¥–ø–∏—Å—å –∏–∑ –≥–æ—Å—É–¥–∞—Ä—Å—Ç–≤–µ–Ω–Ω—ã—Ö –¥–∞–Ω–Ω—ã—Ö.
    caption = state_data.get('caption')
    user_ids = get_user_ids()  # –ü–æ–ª—É—á–∞–µ–º —Å–ø–∏—Å–æ–∫ —É–Ω–∏–∫–∞–ª—å–Ω—ã—Ö ID –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π –∏–∑ –±–∞–∑—ã –¥–∞–Ω–Ω—ã—Ö
    if user_ids:
        for user_id in user_ids:  # –†–∞—Å—Å—ã–ª–∫–∞ –∏–∑–æ–±—Ä–∞–∂–µ–Ω–∏—è —Å –ø–æ–¥–ø–∏—Å—å—é –≤—Å–µ–º –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è–º –∏–∑ —Å–ø–∏—Å–∫–∞
            try:
                await bot.send_photo(user_id, photo, caption=caption)  # –û—Ç–ø—Ä–∞–≤–ª—è–µ–º –∏–∑–æ–±—Ä–∞–∂–µ–Ω–∏–µ —Å –ø–æ–¥–ø–∏—Å—å—é
            except Exception as e:
                print(f"–û—à–∏–±–∫–∞ –ø—Ä–∏ –æ—Ç–ø—Ä–∞–≤–∫–µ –∏–∑–æ–±—Ä–∞–∂–µ–Ω–∏—è —Å –ø–æ–¥–ø–∏—Å—å—é –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—é {user_id}: {str(e)}")
    await message.answer("–ò–∑–æ–±—Ä–∞–∂–µ–Ω–∏–µ —É—Å–ø–µ—à–Ω–æ —Ä–∞–∑–æ—Å–ª–∞–Ω–æ –≤—Å–µ–º –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è–º.")
    await state.clear()


@router.message(Command("send_a_message_to_bot_users"))
async def send_a_message_to_bot_users(message: types.Message, state: FSMContext):
    """–ó–∞–ø—Ä–∞—à–∏–≤–∞–µ—Ç —Ç–µ–∫—Å—Ç —Å–æ–æ–±—â–µ–Ω–∏—è —É –∞–¥–º–∏–Ω–∏—Å—Ç—Ä–∞—Ç–æ—Ä–∞"""
    await state.clear()  # –ó–∞–≤–µ—Ä—à–∞–µ–º —Ç–µ–∫—É—â–µ–µ —Å–æ—Å—Ç–æ—è–Ω–∏–µ –º–∞—à–∏–Ω—ã —Å–æ—Å—Ç–æ—è–Ω–∏–π
    try:
        if message.from_user.id not in [535185511, 301634256]:
            await message.reply('–£ –≤–∞—Å –Ω–µ—Ç –¥–æ—Å—Ç—É–ø–∞ –∫ —ç—Ç–æ–π –∫–æ–º–∞–Ω–¥–µ.')
            return
        await bot.send_message(message.from_user.id, text="–í–≤–µ–¥–∏—Ç–µ —Ç–µ–∫—Å—Ç –¥–ª—è —Ä–∞—Å—Å—ã–ª–∫–∏:")
        await state.set_state(MyStates.waiting_for_message)  # –£—Å—Ç–∞–Ω–∞–≤–ª–∏–≤–∞–µ–º —Å–æ—Å—Ç–æ—è–Ω–∏–µ "–æ–∂–∏–¥–∞–Ω–∏–µ —Å–æ–æ–±—â–µ–Ω–∏—è"
    except Exception as e:
        logger.error(e)


@router.message(StateFilter(MyStates.waiting_for_message))
async def process_send_message(message: types.Message, state: FSMContext):
    """
    –≠—Ç–æ—Ç —Ö–µ–Ω–¥–ª–µ—Ä –±—É–¥–µ—Ç –∂–¥–∞—Ç—å –≤–≤–µ–¥–µ–Ω–Ω–æ–≥–æ —Ç–µ–∫—Å—Ç–∞ –∏ –≤—ã–ø–æ–ª–Ω—è—Ç—å —Ä–∞—Å—Å—ã–ª–∫—É
    """
    # –ü–æ–ª—É—á–∞–µ–º —Å–ø–∏—Å–æ–∫ —É–Ω–∏–∫–∞–ª—å–Ω—ã—Ö ID –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π –∏–∑ –±–∞–∑—ã –¥–∞–Ω–Ω—ã—Ö
    message_text = message.text
    user_ids = get_user_ids()
    if user_ids:
        for user_id in user_ids:  # –†–∞—Å—Å—ã–ª–∫–∞ —Å–æ–æ–±—â–µ–Ω–∏—è –≤—Å–µ–º –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è–º –∏–∑ —Å–ø–∏—Å–∫–∞
            try:
                await bot.send_message(chat_id=user_id, text=message_text, parse_mode="HTML")
            except Exception as e:
                print(f"–û—à–∏–±–∫–∞ –ø—Ä–∏ –æ—Ç–ø—Ä–∞–≤–∫–µ —Å–æ–æ–±—â–µ–Ω–∏—è –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—é {user_id}: {str(e)}")
    await message.answer("–°–æ–æ–±—â–µ–Ω–∏–µ —É—Å–ø–µ—à–Ω–æ —Ä–∞–∑–æ—Å–ª–∞–Ω–æ –≤—Å–µ–º –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è–º.")
    await state.clear()


def get_user_ids():
    """–ü–æ–ª—É—á–∞–µ–º —É–Ω–∏–∫–∞–ª—å–Ω—ã–µ ID –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π –∏–∑ –±–∞–∑—ã –¥–∞–Ω–Ω—ã—Ö"""
    try:
        conn = sqlite3.connect('your_database.db')  # –ó–∞–º–µ–Ω–∏—Ç–µ 'your_database.db' –Ω–∞ –∏–º—è –≤–∞—à–µ–π –±–∞–∑—ã –¥–∞–Ω–Ω—ã—Ö
        cursor = conn.cursor()
        cursor.execute("SELECT DISTINCT user_id FROM users_start")
        user_ids = [row[0] for row in cursor.fetchall()]
        return user_ids
    except Exception as e:
        print(f"–û—à–∏–±–∫–∞ –ø—Ä–∏ –ø–æ–ª—É—á–µ–Ω–∏–∏ ID –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π –∏–∑ –±–∞–∑—ã –¥–∞–Ω–Ω—ã—Ö: {str(e)}")
        return []


def register_admin_greeting_handler():
    """–†–µ–≥–∏—Å—Ç—Ä–∏—Ä—É–µ–º handlers –¥–ª—è –±–æ—Ç–∞"""
    dp.message.register(admin_send_start)
    dp.message.register(export_data)
    dp.message.register(get_users_who_launched_the_bot)
    dp.message.register(send_a_message_to_bot_users)
    dp.message.register(send_an_image_to_bot_users)
