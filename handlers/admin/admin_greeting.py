import os

import openpyxl
from aiogram import types
from aiogram.filters import Command
from aiogram.fsm.context import FSMContext
from aiogram.types import FSInputFile
from loguru import logger

from database.database import reading_from_database
from system.dispatcher import bot, ADMIN_USER_ID
from system.dispatcher import router


@router.message(Command('help'))
async def admin_send_start(message: types.Message, state: FSMContext):
    """Обработчик команды /start, он же пост приветствия 👋"""
    await state.clear()  # Завершаем текущее состояние машины состояний
    """Админ панель"""
    if message.from_user.id not in ADMIN_USER_ID:
        await message.reply("У вас нет прав на выполнение этой команды.")
        return
    await message.answer("Команды админа:\n\n"

                         "<b>Редактирование текста и отправка сообщений:</b>\n\n"

                         "<b>Редактирование текста:</b>\n"
                         "✔️ /edit_types_of_packaging - Виды упаковки\n"
                         "✔️ /edit_services_and_prices - Услуги и цены\n"
                         "✔️ /edit_wooden_corners_bag_tape - Деревянные уголки + мешок + скотч\n"
                         "✔️ /edit_wooden_sheathing_bag_tape - Деревянная обрешетка + мешок + скотч\n"
                         "✔️ /edit_order_form - 🗒 Бланк заказа\n"
                         "✔️ /edit_bag_tape - Мешок + скотч\n"
                         "✔️ /edit_box_bag_tape - Коробка + мешок + скотч\n"
                         "✔️ /edit_pallet_crate - Паллет в обрешетке\n"
                         "✔️ /edit_pallet_with_a_solid_wooden_box - Паллет с глухим деревянным коробом\n"
                         "✔️ /edit_cargo_delivery_prices - Прайсы на доставку Карго\n"
                         "✔️ /edit_goods_redemption_service - Услуга Выкупа товаров\n"
                         "✔️ /edit_product_search_service - Услуга Поиска товаров (производителей в Китае)\n"
                         "✔️ /edit_supplier_inspection - Инспекция поставщиков по провинциям (выезд на производство)\n"
                         "✔️ /edit_wechat_registration_service - Услуга регистрации в WeChat\n"
                         "✔️ /edit_purchase_a_supplier_database - Приобрести базу данных поставщиков\n"
                         "✔️ /edit_what_payments_await_me - Какие платежи меня ожидают?\n"
                         "✔️ /edit_how_is_payment_made - Как совершается оплата?\n"
                         "✔️ /edit_self_redemption - 🛍 Самовыкуп\n"
                         "✔️ /edit_reviews - 💌 Отзывы\n"
                         "✔️ /edit_partnership_conditions_for_intermediaries_button - Партнерские условия для посредников"
                         "✔️ /edit_main_menu - текст меню бота\n"
                         "✔️ /edit_white_cargo_gte - Белая доставка грузов с ГТД\n\n"

                         "<b>Получение данных:</b>\n"
                         "✔️ /get_a_list_of_users_registered_in_the_bot - Получение списка зарегистрированных "
                         "пользователей\n"
                         "✔️ /get_users_who_launched_the_bot - Получение данных пользователей, запускающих бота\n\n"

                         "<b>Отправка сообщений:</b>\n"
                         "✔️ /send_an_image_to_bot_users - Отправка изображения через бота + текст\n"
                         "✔️ /send_a_message_to_bot_users - Отправка текста через бота\n\n"

                         "<b>Замена изображения постов и файлов:</b>\n"
                         "✔️ /greeting_photo - Пост приветствие\n"
                         "✔️ /get_price_lists_file - Замена Прейскуранта CFORB.xlsx"
                         "✔️ /services_and_prices_photo - Услуги и цены\n"
                         "✔️ /white_cargo_gte_photo - Белая доставка грузов с ГТД "
                         "✔️ /types_of_packaging_photo - Виды упаковки\n\n"

                         "/start - начальное меню\n", parse_mode="HTML")


# Функция для создания файла Excel с данными заказов
def create_excel_file(orders):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    # Заголовки столбцов
    sheet['A1'] = 'ID аккаунта пользователя'
    sheet['B1'] = 'Имя'
    sheet['C1'] = 'Фамилия'
    sheet['D1'] = 'Город'
    sheet['E1'] = 'Номер телефона'
    sheet['F1'] = 'Дата регистрации'
    # Заполнение данными заказов
    for index, order in enumerate(orders, start=2):
        sheet.cell(row=index, column=1).value = order[0]  # ID аккаунта пользователя
        sheet.cell(row=index, column=2).value = order[1]  # Имя
        sheet.cell(row=index, column=3).value = order[2]  # Фамилия
        sheet.cell(row=index, column=4).value = order[3]  # Город
        sheet.cell(row=index, column=5).value = order[4]  # Номер телефона
        sheet.cell(row=index, column=6).value = order[5]  # Дата регистрации

    return workbook


@router.message(Command('get_a_list_of_users_registered_in_the_bot'))
async def export_data(message: types.Message, state: FSMContext):
    """Получение списка зарегистрированных пользователей"""
    await state.clear()  # Завершаем текущее состояние машины состояний
    try:
        if message.from_user.id not in [535185511, 301634256]:
            await message.reply('У вас нет доступа к этой команде.')
            return
        # Создание файла Excel
        workbook = create_excel_file(reading_from_database())
        filename = 'Зарегистрированные пользователи в боте.xlsx'
        workbook.save(filename)  # Сохранение файла
        await bot.send_document(message.from_user.id,
                                document=FSInputFile(filename),
                                caption=("Данные пользователей зарегистрированных в боте\n\n"
                                         "Для возврата в начальное меню нажми на /start или /help")
                                )  # Отправка файла пользователю
        os.remove(filename)  # Удаление файла
    except Exception as e:
        logger.error(e)


def create_excel_file_start(orders):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    # Заголовки столбцов
    sheet['A1'] = 'ID аккаунта пользователя'
    sheet['B1'] = 'username'
    sheet['C1'] = 'Имя'
    sheet['D1'] = 'Фамилия'
    sheet['E1'] = 'Дата запуска бота'
    # Заполнение данными заказов
    for index, order in enumerate(orders, start=2):
        sheet.cell(row=index, column=1).value = order[0]  # ID аккаунта пользователя
        sheet.cell(row=index, column=2).value = order[1]  # username
        sheet.cell(row=index, column=3).value = order[2]  # Имя
        sheet.cell(row=index, column=4).value = order[3]  # Фамилия
        sheet.cell(row=index, column=5).value = order[4]  # Дата запуска бота

    return workbook


@router.message(Command("get_users_who_launched_the_bot"))
async def get_users_who_launched_the_bot(message: types.Message, state: FSMContext):
    """Получение данных пользователей, запускающих бота"""
    await state.clear()  # Завершаем текущее состояние машины состояний
    try:
        if message.from_user.id not in [535185511, 301634256]:
            await message.reply('У вас нет доступа к этой команде.')
            return
        workbook = create_excel_file_start(reading_from_database())  # Создание файла Excel
        filename = 'Данные пользователей запустивших бота.xlsx'
        workbook.save(filename)  # Сохранение файла
        await bot.send_document(message.from_user.id, document=FSInputFile(filename),
                                caption=("Данные пользователей зарегистрированных в боте\n\n"
                                         "Для возврата в начальное меню нажми на /start или /help"))  # Отправка файла пользователю
        os.remove(filename)  # Удаление файла
    except Exception as e:
        logger.error(e)


def register_admin_greeting_handler():
    """Регистрируем handlers для бота"""
    router.message.register(admin_send_start)
    router.message.register(export_data)
    router.message.register(get_users_who_launched_the_bot)
